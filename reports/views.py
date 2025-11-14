from django.shortcuts import render, get_object_or_404, redirect
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView, View
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.http import HttpResponse, JsonResponse, Http404
from django.db.models import Count, Sum, Avg, Q
from django.utils import timezone
from django.urls import reverse_lazy, reverse
from django.contrib import messages
from django.utils.translation import gettext as _
from datetime import datetime, timedelta
import csv
import json
from tenants.models import School
from students.models import Student
from accounts.models import User
from academics.models import Class, Section
from django import forms

# Import our report models
from .models import ReportType, SavedReport, ReportDistribution

# Try to import optional libraries
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False


class ReportsIndexView(LoginRequiredMixin, TemplateView):
    """Reports dashboard"""
    template_name = 'reports/index.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['school_slug'] = self.kwargs.get('school_slug', '')
        
        # Get school
        school_slug = self.kwargs.get('school_slug')
        try:
            school = School.objects.get(slug=school_slug)
            context['school'] = school
            
            # Quick stats
            context['total_students'] = Student.objects.filter(current_class__school=school, is_active=True).count()
            context['total_teachers'] = User.objects.filter(role='teacher', is_active=True).count()
            context['total_classes'] = Class.objects.filter(school=school, is_active=True).count()
        except School.DoesNotExist:
            pass
        
        return context


class StudentEnrollmentReportView(LoginRequiredMixin, TemplateView):
    """Student enrollment report"""
    template_name = 'reports/student_enrollment.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        school_slug = self.kwargs.get('school_slug', '')
        context['school_slug'] = school_slug
        
        try:
            school = School.objects.get(slug=school_slug)
            context['school'] = school
            
            # Get enrollment by class
            classes = Class.objects.filter(school=school, is_active=True)
            enrollment_data = []
            
            for cls in classes:
                total = Student.objects.filter(current_class=cls, is_active=True).count()
                male = Student.objects.filter(current_class=cls, is_active=True, gender='male').count()
                female = Student.objects.filter(current_class=cls, is_active=True, gender='female').count()
                
                enrollment_data.append({
                    'class': cls,
                    'total': total,
                    'male': male,
                    'female': female
                })
            
            context['enrollment_data'] = enrollment_data
            context['total_students'] = sum(d['total'] for d in enrollment_data)
            context['total_male'] = sum(d['male'] for d in enrollment_data)
            context['total_female'] = sum(d['female'] for d in enrollment_data)
            
        except School.DoesNotExist:
            context['enrollment_data'] = []
        
        return context


class ExportStudentEnrollmentView(LoginRequiredMixin, View):
    """Export student enrollment report"""
    
    def get(self, request, school_slug, format='csv'):
        school = get_object_or_404(School, slug=school_slug)
        
        # Get data
        classes = Class.objects.filter(school=school, is_active=True)
        data = []
        
        for cls in classes:
            total = Student.objects.filter(current_class=cls, is_active=True).count()
            male = Student.objects.filter(current_class=cls, is_active=True, gender='male').count()
            female = Student.objects.filter(current_class=cls, is_active=True, gender='female').count()
            
            data.append({
                'class': cls.name,
                'total': total,
                'male': male,
                'female': female
            })
        
        if format == 'csv':
            return self.export_csv(data, school)
        elif format == 'excel' and EXCEL_AVAILABLE:
            return self.export_excel(data, school)
        elif format == 'pdf' and PDF_AVAILABLE:
            return self.export_pdf(data, school)
        else:
            return HttpResponse('Format not supported', status=400)
    
    def export_csv(self, data, school):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="enrollment_report_{school.slug}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Student Enrollment Report', school.name])
        writer.writerow([])
        writer.writerow(['Class', 'Total Students', 'Male', 'Female'])
        
        for row in data:
            writer.writerow([row['class'], row['total'], row['male'], row['female']])
        
        writer.writerow([])
        writer.writerow(['Total', sum(d['total'] for d in data), sum(d['male'] for d in data), sum(d['female'] for d in data)])
        
        return response
    
    def export_excel(self, data, school):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Enrollment Report'
        
        # Header
        ws['A1'] = 'Student Enrollment Report'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'] = school.name
        ws['A2'].font = Font(size=12)
        
        # Column headers
        headers = ['Class', 'Total Students', 'Male', 'Female']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Data
        for row_idx, row_data in enumerate(data, 5):
            ws.cell(row=row_idx, column=1, value=row_data['class'])
            ws.cell(row=row_idx, column=2, value=row_data['total'])
            ws.cell(row=row_idx, column=3, value=row_data['male'])
            ws.cell(row=row_idx, column=4, value=row_data['female'])
        
        # Total
        total_row = len(data) + 6
        ws.cell(row=total_row, column=1, value='Total').font = Font(bold=True)
        ws.cell(row=total_row, column=2, value=sum(d['total'] for d in data)).font = Font(bold=True)
        ws.cell(row=total_row, column=3, value=sum(d['male'] for d in data)).font = Font(bold=True)
        ws.cell(row=total_row, column=4, value=sum(d['female'] for d in data)).font = Font(bold=True)
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="enrollment_report_{school.slug}.xlsx"'
        wb.save(response)
        
        return response
    
    def export_pdf(self, data, school):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="enrollment_report_{school.slug}.pdf"'
        
        doc = SimpleDocTemplate(response, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title = Paragraph(f'<b>Student Enrollment Report</b><br/>{school.name}', styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 0.3*inch))
        
        # Table data
        table_data = [['Class', 'Total Students', 'Male', 'Female']]
        for row in data:
            table_data.append([row['class'], str(row['total']), str(row['male']), str(row['female'])])
        
        table_data.append(['Total', str(sum(d['total'] for d in data)), 
                          str(sum(d['male'] for d in data)), 
                          str(sum(d['female'] for d in data))])
        
        # Create table
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, -1), (-1, -1), colors.beige),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        doc.build(elements)
        
        return response


# Additional report views using the new models
class ReportDashboardView(LoginRequiredMixin, TemplateView):
    """Dashboard for reports using the new report models"""
    template_name = 'reports/dashboard.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['school_slug'] = self.kwargs.get('school_slug', '')
        user = self.request.user
        
        # Get report categories
        try:
            context['academic_reports'] = ReportType.objects.filter(module='academic')
            context['attendance_reports'] = ReportType.objects.filter(module='attendance')
            context['finance_reports'] = ReportType.objects.filter(module='finance')
            context['exam_reports'] = ReportType.objects.filter(module='examination')
        except:
            context['academic_reports'] = []
            context['attendance_reports'] = []
            context['finance_reports'] = []
            context['exam_reports'] = []
        
        # Recent reports
        try:
            if user.is_school_admin:
                context['recent_reports'] = SavedReport.objects.filter(
                    school__slug=context['school_slug']
                ).order_by('-created_at')[:5]
            else:
                context['recent_reports'] = SavedReport.objects.filter(
                    Q(created_by=user) | 
                    Q(is_public=True) | 
                    Q(shared_with=user)
                ).distinct().order_by('-created_at')[:5]
        except:
            # Fallback if there are no reports yet
            context['recent_reports'] = []
        
        # Quick stats
        try:
            context['total_report_types'] = ReportType.objects.count()
            context['total_saved_reports'] = SavedReport.objects.filter(
                Q(created_by=user) | 
                Q(is_public=True) | 
                Q(shared_with=user)
            ).distinct().count()
        except:
            # Fallback if models don't exist yet
            context['total_report_types'] = 0
            context['total_saved_reports'] = 0
        
        return context


class ReportTypeListView(LoginRequiredMixin, ListView):
    """List all report types"""
    model = ReportType
    template_name = 'reports/report_type_list.html'
    context_object_name = 'report_types'
    
    def get_queryset(self):
        try:
            queryset = ReportType.objects.all()
            
            # Filter by module if specified
            module = self.request.GET.get('module')
            if module:
                queryset = queryset.filter(module=module)
            
            # Filter by user role
            user_role = getattr(self.request.user, 'role', None)
            if user_role:
                queryset = queryset.filter(available_to_roles__contains=[user_role])
            
            return queryset
        except:
            # Fallback if model doesn't exist yet
            return ReportType.objects.none()
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['school_slug'] = self.kwargs.get('school_slug', '')
        
        # Available modules
        try:
            context['modules'] = ReportType.objects.values_list('module', flat=True).distinct()
        except:
            context['modules'] = []
        
        # Current filter
        context['current_module'] = self.request.GET.get('module', '')
        
        return context


class SavedReportListView(LoginRequiredMixin, ListView):
    """List all saved reports"""
    model = SavedReport
    template_name = 'reports/saved_report_list.html'
    context_object_name = 'reports'
    paginate_by = 20
    
    def get_queryset(self):
        try:
            user = self.request.user
            
            # Get reports visible to user
            queryset = SavedReport.objects.filter(
                Q(created_by=user) | 
                Q(is_public=True) | 
                Q(shared_with=user)
            ).distinct()
            
            # Apply filters
            report_type = self.request.GET.get('type')
            if report_type:
                queryset = queryset.filter(report_type_id=report_type)
            
            status = self.request.GET.get('status')
            if status:
                queryset = queryset.filter(status=status)
            
            # Default ordering
            queryset = queryset.order_by('-created_at')
            
            return queryset
        except:
            # Fallback if model doesn't exist yet
            return SavedReport.objects.none()
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['school_slug'] = self.kwargs.get('school_slug', '')
        
        # Get report types for filter
        try:
            context['report_types'] = ReportType.objects.all()
            context['status_choices'] = SavedReport.STATUS_CHOICES
        except:
            context['report_types'] = []
            context['status_choices'] = []
        
        # Current filters
        context['current_filters'] = {
            'type': self.request.GET.get('type', ''),
            'status': self.request.GET.get('status', ''),
        }
        
        return context


class SavedReportDetailView(LoginRequiredMixin, DetailView):
    """View a saved report"""
    model = SavedReport
    template_name = 'reports/saved_report_detail.html'
    context_object_name = 'report'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['school_slug'] = self.kwargs.get('school_slug', '')
        
        # Convert report data for template
        report = self.object
        try:
            context['report_data'] = report.get_data_as_list()
            context['summary_items'] = report.get_summary_items()
        except AttributeError:
            context['report_data'] = []
            context['summary_items'] = []
        
        # Check user permissions
        user = self.request.user
        context['can_edit'] = user == report.created_by or getattr(user, 'is_school_admin', False)
        context['can_share'] = user == report.created_by or getattr(user, 'is_school_admin', False)
        
        return context

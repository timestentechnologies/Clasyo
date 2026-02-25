from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from django.contrib import messages
from django.shortcuts import redirect
from django.http import JsonResponse, Http404, HttpResponse
from django.db.models import Q
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from .models import Student
from .models import StudentSubject
from accounts.models import User
from core.models import SystemSetting
from core.utils import generate_email, get_school_slug_from_request
from core.utils import get_current_school
from django.db import transaction
import io
import csv
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.comments import Comment
from datetime import datetime


class StudentListView(LoginRequiredMixin, ListView):
    model = Student
    template_name = 'students/list.html'
    context_object_name = 'students'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = super().get_queryset()
        school_slug = self.kwargs.get('school_slug', '')
        from tenants.models import School
        school = School.objects.filter(slug=school_slug, is_active=True).first() if school_slug else None
        # Optional status filter: 'active' or 'inactive'. Default: show all statuses
        status = self.request.GET.get('status', '').strip().lower()
        if school:
            base = queryset.filter(
                Q(school=school) | Q(current_class__school=school) | Q(user__school=school) | Q(parent_user__school=school) | Q(created_by__school=school)
            ).distinct()
        else:
            base = queryset
        if status == 'active':
            base = base.filter(is_active=True)
        elif status == 'inactive':
            base = base.filter(is_active=False)
        return base.order_by('-created_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        school_slug = self.kwargs.get('school_slug', '')
        context['school_slug'] = school_slug
        
        # Add school object to context
        from tenants.models import School
        try:
            school = School.objects.get(slug=school_slug, is_active=True)
            context['school'] = school
        except School.DoesNotExist:
            context['school'] = None
        
        # Add classes and sections for the form (scoped to this school)
        from academics.models import Class, Section
        if context.get('school'):
            context['classes'] = Class.objects.filter(is_active=True, school=context['school'])
            context['sections'] = Section.objects.filter(is_active=True, class_name__school=context['school'])
        else:
            context['classes'] = Class.objects.filter(is_active=True)
            context['sections'] = Section.objects.filter(is_active=True)
        
        # Add dormitories and rooms for the form
        try:
            from dormitory.models import Dormitory, Room
            context['dormitories'] = Dormitory.objects.filter(is_active=True)
            context['rooms'] = Room.objects.filter(is_active=True)
        except ImportError:
            context['dormitories'] = []
            context['rooms'] = []

        try:
            from clubs.models import Club
            if context.get('school'):
                context['clubs'] = Club.objects.filter(is_active=True, school=context['school']).order_by('name')
            else:
                context['clubs'] = Club.objects.filter(is_active=True).order_by('name')
        except ImportError:
            context['clubs'] = []
        
        return context


class StudentSubjectsView(LoginRequiredMixin, DetailView):
    model = Student
    template_name = 'students/subjects.html'
    context_object_name = 'student'

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return super().dispatch(request, *args, **kwargs)
        if request.user.role not in ('admin', 'teacher', 'superadmin'):
            messages.error(request, "Access denied.")
            return redirect('students:detail', school_slug=kwargs.get('school_slug'), pk=kwargs.get('pk'))
        return super().dispatch(request, *args, **kwargs)

    def _get_active_year(self, school):
        from core.models import AcademicYear

        qs = AcademicYear.objects.filter(is_active=True)
        if school:
            qs = qs.filter(school=school)
        return qs.first()

    def _get_available_subjects(self, student, school, active_year):
        from academics.models import Subject, AssignedSubject

        base = Subject.objects.filter(is_active=True)
        if school:
            base = base.filter(school=school)

        if student.current_class_id and active_year:
            assigned_subject_ids = AssignedSubject.objects.filter(
                is_active=True,
                class_name_id=student.current_class_id,
                academic_year=active_year,
            )
            if student.section_id:
                assigned_subject_ids = assigned_subject_ids.filter(Q(section_id=student.section_id) | Q(section__isnull=True))
            assigned_subject_ids = assigned_subject_ids.values_list('subject_id', flat=True)
            subject_qs = base.filter(id__in=assigned_subject_ids)
            if subject_qs.exists():
                return subject_qs.order_by('name')

        return base.order_by('name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['school_slug'] = self.kwargs.get('school_slug', '')

        school = get_current_school(self.request)
        active_year = self._get_active_year(school)
        context['academic_year'] = active_year

        student = self.get_object()
        context['available_subjects'] = self._get_available_subjects(student, school, active_year)

        if active_year:
            enrollments = StudentSubject.objects.filter(
                student=student,
                academic_year=active_year,
                is_active=True,
            ).select_related('subject').order_by('subject__name')
        else:
            enrollments = StudentSubject.objects.none()

        context['enrollments'] = enrollments
        context['selected_subject_ids'] = set(enrollments.values_list('subject_id', flat=True))
        return context

    def post(self, request, *args, **kwargs):
        student = self.get_object()
        school = get_current_school(self.request)
        active_year = self._get_active_year(school)
        if not active_year:
            messages.error(request, "No active academic year found.")
            return redirect('students:subjects', school_slug=kwargs.get('school_slug'), pk=student.pk)

        raw_ids = request.POST.getlist('subjects')
        subject_ids = []
        for sid in raw_ids:
            try:
                subject_ids.append(int(sid))
            except (TypeError, ValueError):
                continue

        from academics.models import Subject

        subjects_qs = Subject.objects.filter(id__in=subject_ids, is_active=True)
        if school:
            subjects_qs = subjects_qs.filter(school=school)
        subjects = list(subjects_qs)

        with transaction.atomic():
            StudentSubject.objects.filter(
                student=student,
                academic_year=active_year,
            ).update(is_active=False)

            for subj in subjects:
                obj, created = StudentSubject.objects.get_or_create(
                    student=student,
                    subject=subj,
                    academic_year=active_year,
                    defaults={
                        'school': school or student.school,
                        'created_by': request.user,
                        'is_active': True,
                    }
                )
                if not created and not obj.is_active:
                    obj.is_active = True
                    if not obj.school_id:
                        obj.school = school or student.school
                    if not obj.created_by_id:
                        obj.created_by = request.user
                    obj.save(update_fields=['is_active', 'school', 'created_by'])

        messages.success(request, "Student subjects updated successfully.")
        return redirect('students:subjects', school_slug=kwargs.get('school_slug'), pk=student.pk)


class StudentDetailView(LoginRequiredMixin, DetailView):
    model = Student
    context_object_name = 'student'
    
    def get_template_names(self):
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest' or 'HTTP_X_REQUESTED_WITH' in self.request.META:
            return ['students/detail_modal.html']
        return ['students/detail.html']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['school_slug'] = self.kwargs.get('school_slug', '')
        
        # Add classes and sections for edit form (scoped to this school)
        from academics.models import Class, Section
        from tenants.models import School
        school_slug = self.kwargs.get('school_slug', '')
        school = School.objects.filter(slug=school_slug, is_active=True).first() if school_slug else None
        if school:
            context['classes'] = Class.objects.filter(is_active=True, school=school)
            context['sections'] = Section.objects.filter(is_active=True, class_name__school=school)
        else:
            context['classes'] = Class.objects.filter(is_active=True)
            context['sections'] = Section.objects.filter(is_active=True)
        
        return context


@method_decorator(csrf_exempt, name='dispatch')
class StudentCreateView(CreateView):
    model = Student
    template_name = 'students/form.html'
    fields = []
    
    def get_success_url(self):
        return f"/school/{self.kwargs.get('school_slug')}/students/"
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['school_slug'] = self.kwargs.get('school_slug', '')
        # Get admission number prefix from settings
        settings = SystemSetting.objects.first()
        prefix = settings.admission_number_prefix if settings else 'STU'
        context['next_admission_number'] = Student.generate_admission_number(prefix)
        return context
    
    def post(self, request, *args, **kwargs):
        try:
            if not request.user.is_authenticated:
                return JsonResponse({'success': False, 'error': 'Not authenticated'})
            
            from datetime import date
            
            # Get or create settings for prefix
            settings = SystemSetting.objects.first()
            prefix = settings.admission_number_prefix if settings else 'STU'
            
            # Generate admission number or use provided
            admission_number = request.POST.get('admission_number')
            if not admission_number or admission_number.strip() == '':
                admission_number = Student.generate_admission_number(prefix)
            
            # Create user account first
            first_name = request.POST.get('first_name')
            last_name = request.POST.get('last_name')
            email = request.POST.get('email', '').strip()
            
            # Auto-generate email if not provided
            if not email:
                school_slug = get_school_slug_from_request(request)
                email = generate_email(first_name, last_name, school_slug, 'student')
            
            # Check if user with this email already exists
            if User.objects.filter(email=email).exists():
                return JsonResponse({
                    'success': False, 
                    'error': f'A user with email {email} already exists. Please use a different email.'
                })
            
            from tenants.models import School
            school_slug = self.kwargs.get('school_slug', '')
            school = School.objects.filter(slug=school_slug, is_active=True).first() if school_slug else None

            user = User.objects.create(
                email=email,
                password='student123',  # Will be hashed
                first_name=first_name,
                last_name=last_name,
                role='student',
                is_active=True
            )
            user.set_password('student123')  # Hash the password properly
            if school:
                user.school = school
            user.save()
            
            # Get admission date or use today
            admission_date = request.POST.get('admission_date')
            if not admission_date:
                admission_date = date.today()
            
            # Create parent user account if father email provided or can be generated
            parent_user = None
            father_email = request.POST.get('father_email', '').strip()
            mother_email = request.POST.get('mother_email', '').strip()
            father_name = request.POST.get('father_name', '').strip()
            mother_name = request.POST.get('mother_name', '').strip()
            
            # Determine parent email and name
            parent_email = father_email if father_email else mother_email
            parent_full_name = father_name if father_name else mother_name
            parent_first_name = parent_full_name.split()[0] if parent_full_name else ''
            parent_last_name = parent_full_name.split()[-1] if parent_full_name and ' ' in parent_full_name else last_name
            
            # Auto-generate parent email if name provided but no email
            if not parent_email and parent_full_name:
                school_slug = get_school_slug_from_request(request)
                parent_email = generate_email(parent_first_name, parent_last_name, school_slug, 'parent')
            
            if parent_email:
                # Check if parent account already exists
                try:
                    parent_user = User.objects.get(email=parent_email, role='parent')
                except User.DoesNotExist:
                    # Create new parent account
                    try:
                        parent_user = User.objects.create(
                            email=parent_email,
                            first_name=parent_first_name or 'Parent',
                            last_name=parent_last_name,  # Already set to student's last name if not provided
                            role='parent',
                            phone=request.POST.get('father_phone', '') or request.POST.get('mother_phone', ''),
                            is_active=True
                        )
                        parent_user.set_password('parent123')  # Default password
                        if parent_user and school and not getattr(parent_user, 'school', None):
                            parent_user.school = school
                            parent_user.save(update_fields=['school'])
                    except Exception as e:
                        print(f"[WARNING] Could not create parent account: {e}")

            # Resolve current class/section scoped to school
            from academics.models import Class, Section
            current_class = None
            section = None
            current_class_id = request.POST.get('current_class')
            if current_class_id:
                if school:
                    current_class = Class.objects.filter(id=current_class_id, school=school).first()
                else:
                    current_class = Class.objects.filter(id=current_class_id).first()
            section_id = request.POST.get('section')
            if section_id:
                if school:
                    section = Section.objects.filter(id=section_id, class_name__school=school).first()
                else:
                    section = Section.objects.filter(id=section_id).first()
            
            # Create student with all fields including parent details
            student = Student.objects.create(
                user=user,
                first_name=first_name,
                last_name=last_name,
                email=email,
                admission_number=admission_number,
                admission_date=admission_date,
                date_of_birth=request.POST.get('date_of_birth') or None,
                gender=request.POST.get('gender', ''),
                blood_group=request.POST.get('blood_group', ''),
                religion=request.POST.get('religion', ''),
                current_address=request.POST.get('current_address', ''),
                city=request.POST.get('city', ''),
                state=request.POST.get('state', ''),
                country=request.POST.get('country', 'US'),
                postal_code=request.POST.get('postal_code', ''),
                # Parent Details
                father_name=request.POST.get('father_name', ''),
                father_phone=request.POST.get('father_phone', ''),
                father_email=father_email,
                father_occupation=request.POST.get('father_occupation', ''),
                mother_name=request.POST.get('mother_name', ''),
                mother_phone=request.POST.get('mother_phone', ''),
                mother_email=mother_email,
                mother_occupation=request.POST.get('mother_occupation', ''),
                guardian_name=request.POST.get('guardian_name', ''),
                guardian_phone=request.POST.get('guardian_phone', ''),
                guardian_email=request.POST.get('guardian_email', ''),
                guardian_relation=request.POST.get('guardian_relation', ''),
                parent_user=parent_user,  # Link to parent user account
                current_class=current_class,
                section=section,
                school=school,
                created_by=request.user
            )

            club_ids = request.POST.getlist('club_ids')
            if club_ids:
                from datetime import date
                try:
                    from clubs.models import Club, ClubMembership
                    clubs_qs = Club.objects.filter(id__in=club_ids, is_active=True)
                    if school:
                        clubs_qs = clubs_qs.filter(school=school)
                    for club in clubs_qs:
                        membership, _ = ClubMembership.objects.get_or_create(
                            student=user,
                            club=club,
                            defaults={
                                'application_reason': 'Added by admin',
                                'parent_consent': True,
                                'status': 'active',
                                'join_date': date.today(),
                            },
                        )
                        if membership.status != 'active':
                            membership.status = 'active'
                            membership.join_date = membership.join_date or date.today()
                            if not membership.application_reason:
                                membership.application_reason = 'Added by admin'
                            membership.save(update_fields=['status', 'join_date', 'application_reason'])
                except Exception:
                    pass
            
            return JsonResponse({
                'success': True, 
                'message': 'Student added successfully',
                'admission_number': admission_number
            })
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({'success': False, 'error': str(e)})


@method_decorator(csrf_exempt, name='dispatch')
class StudentUpdateView(UpdateView):
    model = Student
    template_name = 'students/edit_modal.html'
    fields = []
    
    def get_success_url(self):
        return f"/school/{self.kwargs.get('school_slug')}/students/"
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['school_slug'] = self.kwargs.get('school_slug', '')
        context['student'] = self.get_object()
        from academics.models import Class, Section
        from tenants.models import School
        school_slug = self.kwargs.get('school_slug', '')
        school = School.objects.filter(slug=school_slug, is_active=True).first() if school_slug else None
        if school:
            context['classes'] = Class.objects.filter(is_active=True, school=school)
            context['sections'] = Section.objects.filter(is_active=True, class_name__school=school)
        else:
            context['classes'] = Class.objects.filter(is_active=True)
            context['sections'] = Section.objects.filter(is_active=True)
        
        # Add dormitories and rooms for the form
        try:
            from dormitory.models import Dormitory, Room
            context['dormitories'] = Dormitory.objects.filter(is_active=True)
            context['rooms'] = Room.objects.filter(is_active=True)
        except ImportError:
            context['dormitories'] = []
            context['rooms'] = []

        try:
            from clubs.models import Club, ClubMembership
            clubs_qs = Club.objects.filter(is_active=True)
            if school:
                clubs_qs = clubs_qs.filter(school=school)
            context['clubs'] = clubs_qs.order_by('name')
            selected = set()
            if context['student'] and getattr(context['student'], 'user_id', None):
                memberships_qs = ClubMembership.objects.filter(student_id=context['student'].user_id, status='active')
                if school:
                    memberships_qs = memberships_qs.filter(club__school=school)
                selected = set(memberships_qs.values_list('club_id', flat=True))
            context['selected_club_ids'] = selected
        except ImportError:
            context['clubs'] = []
            context['selected_club_ids'] = set()
        
        return context
    
    def post(self, request, *args, **kwargs):
        try:
            if not request.user.is_authenticated:
                return JsonResponse({'success': False, 'error': 'Not authenticated'})
            
            student = self.get_object()
            
            # Update student fields
            student.first_name = request.POST.get('first_name', student.first_name)
            student.last_name = request.POST.get('last_name', student.last_name)
            student.email = request.POST.get('email', student.email)
            student.date_of_birth = request.POST.get('date_of_birth', student.date_of_birth)
            student.gender = request.POST.get('gender', student.gender)
            student.blood_group = request.POST.get('blood_group', student.blood_group)
            student.current_address = request.POST.get('current_address', student.current_address)
            student.city = request.POST.get('city', student.city)
            student.state = request.POST.get('state', student.state)
            student.postal_code = request.POST.get('postal_code', student.postal_code)
            student.admission_date = request.POST.get('admission_date', student.admission_date)
            
            # Update transport and dormitory details
            student.is_transport_required = request.POST.get('is_transport_required') == 'True'
            student.is_hostel_required = request.POST.get('is_hostel_required') == 'True'
            
            if student.is_hostel_required:
                dormitory_id = request.POST.get('dormitory')
                room_id = request.POST.get('room')
                if dormitory_id:
                    try:
                        from dormitory.models import Dormitory
                        student.dormitory = Dormitory.objects.get(id=dormitory_id)
                    except:
                        pass
                if room_id:
                    try:
                        from dormitory.models import Room
                        student.room = Room.objects.get(id=room_id)
                    except:
                        pass
            else:
                student.dormitory = None
                student.room = None
            
            # Update academic details
            current_class_id = request.POST.get('current_class')
            if current_class_id:
                from academics.models import Class
                from tenants.models import School
                school_slug = self.kwargs.get('school_slug', '')
                school = School.objects.filter(slug=school_slug, is_active=True).first() if school_slug else None
                if school:
                    current_class = Class.objects.filter(id=current_class_id, school=school).first()
                else:
                    current_class = Class.objects.filter(id=current_class_id).first()
                if current_class:
                    student.current_class = current_class
                else:
                    student.current_class = None
            
            section_id = request.POST.get('section')
            if section_id:
                from academics.models import Section
                from tenants.models import School
                school_slug = self.kwargs.get('school_slug', '')
                school = School.objects.filter(slug=school_slug, is_active=True).first() if school_slug else None
                if school:
                    section = Section.objects.filter(id=section_id, class_name__school=school).first()
                else:
                    section = Section.objects.filter(id=section_id).first()
                if section:
                    student.section = section
                else:
                    student.section = None
            
            # Ensure student is linked to a school if missing
            if not getattr(student, 'school', None):
                try:
                    if student.current_class and getattr(student.current_class, 'school', None):
                        student.school = student.current_class.school
                    elif student.user and getattr(student.user, 'school', None):
                        student.school = student.user.school
                except Exception:
                    pass
            
            # Update parent details
            student.father_name = request.POST.get('father_name', student.father_name)
            student.father_phone = request.POST.get('father_phone', student.father_phone)
            father_email = request.POST.get('father_email', '').strip()
            student.father_email = father_email
            student.father_occupation = request.POST.get('father_occupation', student.father_occupation)
            student.mother_name = request.POST.get('mother_name', student.mother_name)
            student.mother_phone = request.POST.get('mother_phone', student.mother_phone)
            mother_email = request.POST.get('mother_email', '').strip()
            student.mother_email = mother_email
            student.mother_occupation = request.POST.get('mother_occupation', student.mother_occupation)
            student.guardian_name = request.POST.get('guardian_name', student.guardian_name)
            student.guardian_phone = request.POST.get('guardian_phone', student.guardian_phone)
            student.guardian_email = request.POST.get('guardian_email', student.guardian_email)
            student.guardian_relation = request.POST.get('guardian_relation', student.guardian_relation)
            
            # Create or link parent user account if email provided or can be generated
            parent_email = father_email if father_email else mother_email
            father_name = request.POST.get('father_name', '').strip()
            mother_name = request.POST.get('mother_name', '').strip()
            parent_full_name = father_name if father_name else mother_name
            parent_first_name = parent_full_name.split()[0] if parent_full_name else ''
            parent_last_name = parent_full_name.split()[-1] if parent_full_name and ' ' in parent_full_name else student.last_name
            
            # Auto-generate parent email if name provided but no email
            if not parent_email and parent_full_name and not student.parent_user:
                school_slug = get_school_slug_from_request(request)
                parent_email = generate_email(parent_first_name, parent_last_name, school_slug, 'parent')
                # Update the student's email fields with the auto-generated email
                if father_name:
                    student.father_email = parent_email
                else:
                    student.mother_email = parent_email
            
            if parent_email and not student.parent_user:
                # Check if parent account already exists
                try:
                    parent_user = User.objects.get(email=parent_email, role='parent')
                    student.parent_user = parent_user
                except User.DoesNotExist:
                    # Create new parent account
                    try:
                        parent_user = User.objects.create(
                            email=parent_email,
                            first_name=parent_first_name or 'Parent',
                            last_name=parent_last_name or student.last_name,
                            role='parent',
                            phone=request.POST.get('father_phone', '') or request.POST.get('mother_phone', ''),
                            is_active=True
                        )
                        parent_user.set_password('parent123')  # Default password
                        parent_user.save()
                        student.parent_user = parent_user
                    except Exception as e:
                        print(f"[WARNING] Could not create parent account: {e}")
            
            student.save()

            club_ids = set(map(int, request.POST.getlist('club_ids')))
            if club_ids is not None:
                from datetime import date
                try:
                    from clubs.models import Club, ClubMembership
                    school = get_current_school(request)
                    allowed_clubs = Club.objects.filter(is_active=True)
                    if school:
                        allowed_clubs = allowed_clubs.filter(school=school)
                    allowed_ids = set(allowed_clubs.values_list('id', flat=True))
                    club_ids = set([cid for cid in club_ids if cid in allowed_ids])

                    memberships_qs = ClubMembership.objects.filter(student=student.user, club__in=allowed_clubs)
                    memberships_qs.exclude(club_id__in=club_ids).exclude(status='inactive').update(status='inactive')

                    for club in allowed_clubs.filter(id__in=club_ids):
                        membership, _ = ClubMembership.objects.get_or_create(
                            student=student.user,
                            club=club,
                            defaults={
                                'application_reason': 'Added by admin',
                                'parent_consent': True,
                                'status': 'active',
                                'join_date': date.today(),
                            },
                        )
                        if membership.status != 'active':
                            membership.status = 'active'
                            membership.join_date = membership.join_date or date.today()
                            if not membership.application_reason:
                                membership.application_reason = 'Added by admin'
                            membership.save(update_fields=['status', 'join_date', 'application_reason'])
                except Exception:
                    pass
            
            # Update user email if changed
            if student.user.email != student.email:
                student.user.email = student.email
                student.user.first_name = student.first_name
                student.user.last_name = student.last_name
                student.user.save()
            
            return JsonResponse({'success': True, 'message': 'Student updated successfully'})
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({'success': False, 'error': str(e)})


@method_decorator(csrf_exempt, name='dispatch')
class StudentDeleteView(DeleteView):
    model = Student
    
    def get_success_url(self):
        return f"/school/{self.kwargs.get('school_slug')}/students/"
    
    def post(self, request, *args, **kwargs):
        try:
            if not request.user.is_authenticated:
                return JsonResponse({'success': False, 'error': 'Not authenticated'})
                
            student = self.get_object()
            student.user.delete()  # This will cascade delete the student
            return JsonResponse({'success': True, 'message': 'Student deleted successfully'})
        except Exception as e:
            import traceback
            traceback.print_exc()
            return JsonResponse({'success': False, 'error': str(e)})


class ParentListView(LoginRequiredMixin, ListView):
    model = User
    template_name = 'students/parents.html'
    context_object_name = 'parents'
    paginate_by = 20
    
    def get_queryset(self):
        school_slug = self.kwargs.get('school_slug', '')
        from tenants.models import School
        school = School.objects.filter(slug=school_slug, is_active=True).first() if school_slug else None
        qs = User.objects.filter(role='parent').order_by('-created_at')
        if school:
            qs = qs.filter(
                Q(school=school) |
                Q(children__current_class__school=school) |
                Q(children__user__school=school)
            ).distinct()
        return qs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['school_slug'] = self.kwargs.get('school_slug', '')
        return context


class ParentDetailView(LoginRequiredMixin, DetailView):
    model = User
    
    def get(self, request, *args, **kwargs):
        try:
            parent = self.get_object()
            if parent.role != 'parent':
                return JsonResponse({'success': False, 'error': 'Not a parent user'})
            
            # Get children (students linked to this parent) scoped to this school
            school_slug = self.kwargs.get('school_slug', '')
            from tenants.models import School
            school = School.objects.filter(slug=school_slug, is_active=True).first() if school_slug else None
            children = Student.objects.filter(parent_user=parent)
            if school:
                children = children.filter(Q(current_class__school=school) | Q(user__school=school)).distinct()
            children_data = [{
                'name': f"{child.first_name} {child.last_name}",
                'admission_number': child.admission_number
            } for child in children]
            
            return JsonResponse({
                'success': True,
                'parent': {
                    'name': parent.get_full_name(),
                    'first_name': parent.first_name,
                    'last_name': parent.last_name,
                    'email': parent.email,
                    'phone': str(parent.phone) if parent.phone else '',
                    'is_active': parent.is_active
                },
                'children': children_data
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})


class StudentExportView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        fmt = request.GET.get('format', 'csv').lower()
        school_slug = kwargs.get('school_slug', '')
        from tenants.models import School
        school = School.objects.filter(slug=school_slug, is_active=True).first() if school_slug else None
        students = Student.objects.select_related('current_class', 'section')
        if school:
            students = students.filter(Q(current_class__school=school) | Q(user__school=school) | Q(parent_user__school=school) | Q(created_by__school=school)).distinct()
        else:
            students = students.all()
        headers = [
            'admission_number', 'first_name', 'last_name', 'gender', 'date_of_birth', 'admission_date',
            'current_class', 'section', 'email', 'current_address', 'city', 'state', 'postal_code',
            'father_name', 'father_phone', 'father_email', 'father_occupation',
            'mother_name', 'mother_phone', 'mother_email', 'mother_occupation',
            'guardian_name', 'guardian_phone', 'guardian_email', 'guardian_relation'
        ]
        rows = []
        for s in students:
            rows.append([
                s.admission_number or '', s.first_name or '', s.last_name or '', s.gender or '',
                s.date_of_birth.isoformat() if s.date_of_birth else '',
                s.admission_date.isoformat() if s.admission_date else '',
                getattr(s.current_class, 'name', '') or '',
                getattr(s.section, 'name', '') or '',
                s.email or '', s.current_address or '', s.city or '', s.state or '', s.postal_code or '',
                s.father_name or '', str(s.father_phone) if s.father_phone else '', s.father_email or '', s.father_occupation or '',
                s.mother_name or '', str(s.mother_phone) if s.mother_phone else '', s.mother_email or '', s.mother_occupation or '',
                s.guardian_name or '', str(s.guardian_phone) if s.guardian_phone else '', s.guardian_email or '', s.guardian_relation or ''
            ])
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        if fmt == 'excel' or fmt == 'xlsx':
            wb = Workbook()
            ws = wb.active
            ws.title = 'Students'
            ws.append(headers)
            for r in rows:
                ws.append(r)
            bio = io.BytesIO()
            wb.save(bio)
            bio.seek(0)
            resp = HttpResponse(bio.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            resp['Content-Disposition'] = f'attachment; filename="students_{ts}.xlsx"'
            return resp
        else:
            sio = io.StringIO()
            writer = csv.writer(sio)
            writer.writerow(headers)
            for r in rows:
                writer.writerow(r)
            resp = HttpResponse(sio.getvalue(), content_type='text/csv')
            resp['Content-Disposition'] = f'attachment; filename="students_{ts}.csv"'
            return resp


class StudentTemplateView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        wb = Workbook()
        ws_template = wb.active
        ws_template.title = 'Template'
        headers = [
            'admission_number', 'first_name', 'last_name', 'gender', 'date_of_birth', 'admission_date',
            'current_class', 'section', 'email', 'current_address', 'city', 'state', 'postal_code',
            'father_name', 'father_phone', 'father_email', 'father_occupation',
            'mother_name', 'mother_phone', 'mother_email', 'mother_occupation',
            'guardian_name', 'guardian_phone', 'guardian_email', 'guardian_relation'
        ]
        ws_template.append(headers)
        required_fields = {'first_name', 'last_name', 'gender', 'date_of_birth', 'admission_date', 'current_class', 'current_address', 'city', 'state'}
        autogenerated_fields = {'admission_number', 'email', 'father_email', 'mother_email'}
        required_fill = PatternFill(fill_type='solid', start_color='FFFFF3CC', end_color='FFFFF3CC')
        autogenerated_fill = PatternFill(fill_type='solid', start_color='FFC6EFCE', end_color='FFC6EFCE')
        header_font = Font(bold=True)
        header_align = Alignment(horizontal='center', vertical='center')
        for col_idx, title in enumerate(headers, start=1):
            cell = ws_template.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.alignment = header_align
            if title in required_fields:
                cell.fill = required_fill
                cell.comment = Comment('Required field', 'System')
            elif title in autogenerated_fields:
                cell.fill = autogenerated_fill
                cell.comment = Comment('Auto-generated if left blank', 'System')
        ws_template.freeze_panes = 'A2'
        for col_idx, title in enumerate(headers, start=1):
            ws_template.column_dimensions[ws_template.cell(row=1, column=col_idx).column_letter].width = max(14, len(title) + 2)
        ws_sample = wb.create_sheet('Sample Data')
        ws_sample.append(headers)
        for col_idx, title in enumerate(headers, start=1):
            c = ws_sample.cell(row=1, column=col_idx)
            c.font = header_font
            c.alignment = header_align
            if title in required_fields:
                c.fill = required_fill
            elif title in autogenerated_fields:
                c.fill = autogenerated_fill
        ws_sample.append([
            'STU00001', 'John', 'Doe', 'male', '2012-05-10', '2024-01-10', 'Grade 1', 'A', 'john.doe@school.com',
            '123 Main St', 'Nairobi', 'Nairobi', '00100',
            'James Doe', '+254700000001', 'james.doe@example.com', 'Engineer',
            'Jane Doe', '+254700000002', 'jane.doe@example.com', 'Doctor',
            '', '', '', ''
        ])
        ws_instr = wb.create_sheet('Instructions')
        instructions = [
            ['Legend:'],
            ['Yellow = Required field'],
            ['Green = Auto-generated if left blank'],
            [''],
            ['Fill data in the Template sheet. Dates must be in YYYY-MM-DD format.'],
            ['current_class and section should match existing names in your system.'],
            ['Required fields: first_name, last_name, date_of_birth, gender, admission_date, current_class, current_address, city, state.'],
            ['Auto-generated fields: admission_number, email. Parent emails (father_email/mother_email) will be auto-generated if names are provided and emails are left blank.'],
            ['Phones should include country code, e.g., +2547xxxxxxx.']
        ]
        for row in instructions:
            ws_instr.append(row)
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        resp = HttpResponse(bio.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = f'attachment; filename="students_template_{ts}.xlsx"'
        return resp


@method_decorator(csrf_exempt, name='dispatch')
class StudentImportSheetsView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        f = request.FILES.get('file')
        if not f:
            return JsonResponse({'success': False, 'error': 'No file uploaded'})
        name = f.name.lower()
        if not name.endswith('.xlsx'):
            return JsonResponse({'success': True, 'sheets': []})
        try:
            wb = load_workbook(f, read_only=True)
            return JsonResponse({'success': True, 'sheets': wb.sheetnames})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})


@method_decorator(csrf_exempt, name='dispatch')
class StudentImportPreviewView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        f = request.FILES.get('file')
        if not f:
            return JsonResponse({'success': False, 'error': 'No file uploaded'})
        name = f.name.lower()
        try:
            if name.endswith('.csv'):
                df = pd.read_csv(f, dtype=str)
            elif name.endswith('.xlsx'):
                sheet = request.POST.get('sheet_name')
                df = pd.read_excel(f, sheet_name=sheet if sheet else 0, dtype=str)
            else:
                return JsonResponse({'success': False, 'error': 'Unsupported file type'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Failed to read file: {e}'})

        # Normalize columns
        original_columns = list(df.columns)
        norm_columns = [str(c).strip().lower().replace('\n', ' ').replace('  ', ' ').replace(' ', '_') for c in original_columns]
        df.columns = norm_columns

        # Canonical fields and metadata
        canonical_fields = [
            'admission_number', 'first_name', 'last_name', 'gender', 'date_of_birth', 'admission_date',
            'current_class', 'section', 'email', 'current_address', 'city', 'state', 'postal_code',
            'father_name', 'father_phone', 'father_email', 'father_occupation',
            'mother_name', 'mother_phone', 'mother_email', 'mother_occupation',
            'guardian_name', 'guardian_phone', 'guardian_email', 'guardian_relation'
        ]
        required = {'first_name', 'last_name', 'gender', 'date_of_birth', 'admission_date', 'current_class', 'current_address', 'city', 'state'}
        autogenerated = {'admission_number', 'email', 'father_email', 'mother_email'}

        # Synonyms mapping (normalized -> canonical)
        synonyms = {
            'firstname': 'first_name',
            'first': 'first_name',
            'lastname': 'last_name',
            'last': 'last_name',
            'dob': 'date_of_birth',
            'dateofbirth': 'date_of_birth',
            'admissiondate': 'admission_date',
            'class': 'current_class',
            'class_name': 'current_class',
            'grade': 'current_class',
            'section_name': 'section',
            'zip': 'postal_code',
            'postcode': 'postal_code',
            'father_mobile': 'father_phone',
            'father_tel': 'father_phone',
            'mother_mobile': 'mother_phone',
            'mother_tel': 'mother_phone',
        }

        # Build column mapping
        column_map = {}
        mapped_canonicals = set()
        for src in norm_columns:
            if src in canonical_fields:
                column_map[src] = src
                mapped_canonicals.add(src)
            elif src in synonyms:
                column_map[src] = synonyms[src]
                mapped_canonicals.add(synonyms[src])
            else:
                column_map[src] = None

        missing_required = sorted(list(required - mapped_canonicals))

        # Prepare a working dataframe with canonical columns only
        work_df = pd.DataFrame()
        for src, dst in column_map.items():
            if dst:
                work_df[dst] = df[src]
        work_df = work_df.fillna('')

        total_rows = len(work_df.index)
        row_errors = []

        # Vectorized date validation
        def is_valid_date(series):
            try:
                conv = pd.to_datetime(series, errors='coerce')
                return ~conv.isna()
            except Exception:
                return pd.Series([False] * len(series))

        # Preload reference names for existence checks (prefer school-scoped)
        from academics.models import Class, Section
        school_slug = kwargs.get('school_slug', '')
        from tenants.models import School
        school = School.objects.filter(slug=school_slug, is_active=True).first() if school_slug else None
        if school:
            class_names = set([n.lower() for n in Class.objects.filter(school=school).values_list('name', flat=True)])
            section_names = set([n.lower() for n in Section.objects.filter(class_name__school=school).values_list('name', flat=True)])
        else:
            class_names = set([n.lower() for n in Class.objects.values_list('name', flat=True)])
            section_names = set([n.lower() for n in Section.objects.values_list('name', flat=True)])

        # Validate rows
        for idx in range(total_rows):
            errs = []
            row = work_df.iloc[idx]
            # Required non-empty
            for col in required:
                if col not in work_df.columns or not str(row.get(col, '')).strip():
                    errs.append(f"Missing required: {col}")
            # Dates
            if 'date_of_birth' in work_df.columns and str(row.get('date_of_birth', '')).strip():
                if not is_valid_date(pd.Series([row['date_of_birth']])).iloc[0]:
                    errs.append('Invalid date format: date_of_birth')
            if 'admission_date' in work_df.columns and str(row.get('admission_date', '')).strip():
                if not is_valid_date(pd.Series([row['admission_date']])).iloc[0]:
                    errs.append('Invalid date format: admission_date')
            # Existence checks (best-effort)
            if 'current_class' in work_df.columns and str(row.get('current_class', '')).strip():
                if row['current_class'].strip().lower() not in class_names:
                    errs.append('Unknown class name: current_class')
            if 'section' in work_df.columns and str(row.get('section', '')).strip():
                if row['section'].strip().lower() not in section_names:
                    errs.append('Unknown section name: section')
            if errs:
                row_errors.append({'row': idx + 2, 'errors': errs})

        invalid_rows = len(row_errors)
        can_import = (invalid_rows == 0) and (len(missing_required) == 0)

        # Build first non-empty values per source column
        first_values = {}
        for src in norm_columns:
            val = ''
            try:
                ser = df[src]
                for v in ser:
                    if pd.isna(v):
                        continue
                    s = str(v).strip()
                    if s:
                        val = s
                        break
            except Exception:
                val = ''
            first_values[src] = val

        # Build columns info
        columns_info = []
        for original, norm in zip(original_columns, norm_columns):
            mapped = column_map.get(norm)
            columns_info.append({
                'source': original,
                'normalized': norm,
                'mapped_to': mapped or '',
                'first_value': first_values.get(norm, ''),
                'is_required': (mapped in required) if mapped else False,
                'is_autogenerated': (mapped in autogenerated) if mapped else False,
                'status': 'mapped' if mapped else 'unmapped'
            })

        # Predict potential creates/updates
        potential = {'updates': 0, 'creates': 0, 'unknown': 0}
        try:
            from students.models import Student
            adms = set()
            emails = set()
            if 'admission_number' in work_df.columns:
                adms = set([str(x).strip() for x in work_df['admission_number'].tolist() if str(x).strip()])
            if 'email' in work_df.columns:
                emails = set([str(x).strip() for x in work_df['email'].tolist() if str(x).strip()])
            existing_adm = set()
            existing_email = set()
            if adms:
                existing_adm = set(Student.objects.filter(admission_number__in=list(adms)).values_list('admission_number', flat=True))
            if emails:
                existing_email = set(Student.objects.filter(email__in=list(emails)).values_list('email', flat=True))
            for idx in range(total_rows):
                adm = str(work_df.iloc[idx].get('admission_number', '')).strip() if 'admission_number' in work_df.columns else ''
                eml = str(work_df.iloc[idx].get('email', '')).strip() if 'email' in work_df.columns else ''
                if (adm and adm in existing_adm) or (eml and eml in existing_email):
                    potential['updates'] += 1
                elif adm or eml:
                    potential['creates'] += 1
                else:
                    potential['unknown'] += 1
        except Exception:
            pass

        preview = {
            'success': True,
            'columns': columns_info,
            'missing_required': missing_required,
            'stats': {
                'total_rows': int(total_rows),
                'invalid_rows': int(invalid_rows),
                'valid_rows': int(total_rows - invalid_rows)
            },
            'sample_errors': row_errors[:50],
            'can_import': can_import,
            'potential_actions': potential
        }
        return JsonResponse(preview)

@method_decorator(csrf_exempt, name='dispatch')
class StudentImportView(LoginRequiredMixin, View):
    def post(self, request, *args, **kwargs):
        f = request.FILES.get('file')
        if not f:
            return JsonResponse({'success': False, 'error': 'No file uploaded'})
        name = f.name.lower()
        try:
            if name.endswith('.csv'):
                df = pd.read_csv(f, dtype=str)
            elif name.endswith('.xlsx'):
                sheet = request.POST.get('sheet_name')
                df = pd.read_excel(f, sheet_name=sheet if sheet else 0, dtype=str)
            else:
                return JsonResponse({'success': False, 'error': 'Unsupported file type'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Failed to read file: {e}'})
        df.columns = [str(c).strip().lower().replace(' ', '_') for c in df.columns]
        df = df.fillna('')
        required = ['first_name', 'last_name', 'date_of_birth', 'gender', 'admission_date', 'current_class', 'current_address', 'city', 'state']
        from academics.models import Class, Section
        from tenants.models import School
        school_slug = kwargs.get('school_slug', '')
        school = School.objects.filter(slug=school_slug, is_active=True).first() if school_slug else None
        success_count = 0
        created_count = 0
        updated_count = 0
        errors = []
        for idx, row in enumerate(df.to_dict(orient='records')):
            try:
                for col in required:
                    if not str(row.get(col, '')).strip():
                        raise ValueError(f'Missing required field: {col}')
                dob = pd.to_datetime(row.get('date_of_birth'), errors='coerce')
                adm_dt = pd.to_datetime(row.get('admission_date'), errors='coerce')
                if pd.isna(dob) or pd.isna(adm_dt):
                    raise ValueError('Invalid date format, use YYYY-MM-DD')
                dob = dob.date()
                adm_dt = adm_dt.date()
                class_name = str(row.get('current_class', '')).strip()
                section_name = str(row.get('section', '')).strip()
                if school:
                    current_class = Class.objects.filter(name__iexact=class_name, school=school).first() if class_name else None
                    section = Section.objects.filter(name__iexact=section_name, class_name__school=school).first() if section_name else None
                else:
                    current_class = Class.objects.filter(name__iexact=class_name).first() if class_name else None
                    section = Section.objects.filter(name__iexact=section_name).first() if section_name else None
                admission_number = str(row.get('admission_number', '')).strip() or None
                email = str(row.get('email', '')).strip()
                first_name = str(row.get('first_name', '')).strip()
                last_name = str(row.get('last_name', '')).strip()
                if not email:
                    school_slug = get_school_slug_from_request(request)
                    email = generate_email(first_name, last_name, school_slug, 'student')
                student = None
                if admission_number:
                    student = Student.objects.filter(admission_number=admission_number).first()
                if not student and email:
                    student = Student.objects.filter(email=email).first()
                if student and student.user:
                    user = student.user
                else:
                    user = User.objects.filter(email=email, role='student').first()
                    if not user:
                        user = User.objects.create(email=email, first_name=first_name, last_name=last_name, role='student', is_active=True)
                        user.set_password('student123')
                        if school:
                            user.school = school
                        user.save()
                # Ensure imported student's user is linked to this school
                if school and not getattr(user, 'school', None):
                    user.school = school
                    user.save(update_fields=['school'])
                parent_user = None
                father_email = str(row.get('father_email', '')).strip()
                mother_email = str(row.get('mother_email', '')).strip()
                father_name = str(row.get('father_name', '')).strip()
                mother_name = str(row.get('mother_name', '')).strip()
                parent_email = father_email if father_email else mother_email
                parent_full_name = father_name if father_name else mother_name
                parent_first_name = parent_full_name.split()[0] if parent_full_name else ''
                parent_last_name = parent_full_name.split()[-1] if parent_full_name and ' ' in parent_full_name else last_name
                if not parent_email and parent_full_name:
                    school_slug = get_school_slug_from_request(request)
                    parent_email = generate_email(parent_first_name, parent_last_name, school_slug, 'parent')
                if parent_email:
                    parent_user = User.objects.filter(email=parent_email, role='parent').first()
                    if not parent_user:
                        parent_user = User.objects.create(email=parent_email, first_name=parent_first_name or 'Parent', last_name=parent_last_name or last_name, role='parent', phone=row.get('father_phone', '') or row.get('mother_phone', ''), is_active=True)
                        parent_user.set_password('parent123')
                        if school:
                            parent_user.school = school
                        parent_user.save()
                # Ensure imported parent user is linked to this school
                if parent_user and school and not getattr(parent_user, 'school', None):
                    parent_user.school = school
                    parent_user.save(update_fields=['school'])
                data = {
                    'user': user,
                    'first_name': first_name,
                    'last_name': last_name,
                    'email': email,
                    'admission_number': admission_number or Student.generate_admission_number('STU'),
                    'admission_date': adm_dt,
                    'date_of_birth': dob,
                    'gender': str(row.get('gender', '')).strip(),
                    'blood_group': str(row.get('blood_group', '')).strip(),
                    'religion': str(row.get('religion', '')).strip(),
                    'current_address': str(row.get('current_address', '')).strip(),
                    'city': str(row.get('city', '')).strip(),
                    'state': str(row.get('state', '')).strip(),
                    'country': str(row.get('country', 'US')).strip() or 'US',
                    'postal_code': str(row.get('postal_code', '')).strip(),
                    'father_name': father_name,
                    'father_phone': row.get('father_phone', ''),
                    'father_email': father_email,
                    'father_occupation': str(row.get('father_occupation', '')).strip(),
                    'mother_name': mother_name,
                    'mother_phone': row.get('mother_phone', ''),
                    'mother_email': mother_email,
                    'mother_occupation': str(row.get('mother_occupation', '')).strip(),
                    'guardian_name': str(row.get('guardian_name', '')).strip(),
                    'guardian_phone': row.get('guardian_phone', ''),
                    'guardian_email': str(row.get('guardian_email', '')).strip(),
                    'guardian_relation': str(row.get('guardian_relation', '')).strip(),
                    'current_class': current_class,
                    'section': section,
                    'parent_user': parent_user,
                    'created_by': request.user
                }
                with transaction.atomic():
                    if student:
                        for k, v in data.items():
                            setattr(student, k, v)
                        # Backfill school on existing student if missing
                        if school and not getattr(student, 'school', None):
                            student.school = school
                        student.save()
                        updated_count += 1
                    else:
                        # Ensure new student is linked to this school
                        student = Student.objects.create(**data, school=school)
                        created_count += 1
                success_count += 1
            except Exception as e:
                errors.append({'row': idx + 2, 'error': str(e)})
        return JsonResponse({'success': True, 'imported': success_count, 'created': created_count, 'updated': updated_count, 'errors': len(errors), 'error_rows': errors})
